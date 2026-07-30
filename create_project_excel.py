import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import csv
import os

def create_project_data_sheet():
    wb = openpyxl.Workbook()
    
    # ---------------------------------------------------------
    # SHEET 1: Project Content Collection Form
    # ---------------------------------------------------------
    ws_form = wb.active
    ws_form.title = "Project_Content_Form"
    ws_form.views.sheetView[0].showGridLines = True

    # Color Palette - Professional Corporate Gold & Slate Navy
    NAVY_HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid") # Dark Charcoal/Navy
    GOLD_BANNER_FILL = PatternFill(start_color="C5A059", end_color="C5A059", fill_type="solid") # Muted Luxury Gold
    SUBHEADER_FILL = PatternFill(start_color="374151", end_color="374151", fill_type="solid")
    ZEBRA_FILL = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
    WHITE_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    # Fonts
    TITLE_FONT = Font(name="Segoe UI", size=16, bold=True, color="FFFFFF")
    SUBTITLE_FONT = Font(name="Segoe UI", size=10, italic=True, color="F3F4F6")
    BANNER_FONT = Font(name="Segoe UI", size=11, bold=True, color="1F2937")
    HEADER_FONT = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    DATA_FONT = Font(name="Segoe UI", size=10, color="1F2937")
    SAMPLE_DATA_FONT = Font(name="Segoe UI", size=9, italic=True, color="4B5563")

    # Borders
    THIN_BORDER_SIDE = Side(border_style="thin", color="E5E7EB")
    BORDER_ALL = Border(left=THIN_BORDER_SIDE, right=THIN_BORDER_SIDE, top=THIN_BORDER_SIDE, bottom=THIN_BORDER_SIDE)
    HEADER_BORDER = Border(left=THIN_BORDER_SIDE, right=THIN_BORDER_SIDE, top=Side(border_style="medium", color="1F2937"), bottom=Side(border_style="medium", color="1F2937"))

    # Title Block
    ws_form.merge_cells("A1:P1")
    title_cell = ws_form["A1"]
    title_cell.value = "BNP INTERIORS — CLIENT PROJECT CONTENT COLLECTION FORM"
    title_cell.font = TITLE_FONT
    title_cell.fill = NAVY_HEADER_FILL
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_form.row_dimensions[1].height = 35

    ws_form.merge_cells("A2:P2")
    sub_cell = ws_form["A2"]
    sub_cell.value = "Please fill in the project details below to populate your website's portfolio detail pages. For support or questions, contact your Project Lead."
    sub_cell.font = SUBTITLE_FONT
    sub_cell.fill = NAVY_HEADER_FILL
    sub_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_form.row_dimensions[2].height = 22

    # Instructions Banner
    ws_form.merge_cells("A3:P3")
    banner_cell = ws_form["A3"]
    banner_cell.value = "📌 Note: Columns marked with (*) are mandatory. Please paste high-resolution photo folder links in the LAST column (Column P)."
    banner_cell.font = BANNER_FONT
    banner_cell.fill = GOLD_BANNER_FILL
    banner_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws_form.row_dimensions[3].height = 25

    # Blank gap row 4
    ws_form.row_dimensions[4].height = 10

    # Column Headers (Row 5)
    headers = [
        ("S.No.", 8),
        ("Project Name / Title*", 24),
        ("Category / Sector*", 20),
        ("Client / Organization Name", 24),
        ("City & State / Location*", 22),
        ("Project Scope / Services Provided*", 30),
        ("Built-up Area / Scale (Sq. Ft. / Sq. Mtr)*", 25),
        ("Capacity / Quantum Details (Rooms, Key Elements)", 30),
        ("Project Status / Completion Year*", 22),
        ("Short Summary / Overview (2-3 Sentences)*", 40),
        ("Detailed Design Concept & Story*", 50),
        ("Key Design Highlights & Features (Bullet Points)", 40),
        ("Key Materials & Finishes Used", 35),
        ("Architect / Lead Designer Credits", 26),
        ("Client Testimonial / Quote (Optional)", 35),
        ("Drive Link of Image Folder*", 45) # LAST COLUMN as requested
    ]

    ws_form.row_dimensions[5].height = 32

    raw_headers = []
    for col_idx, (header_text, col_width) in enumerate(headers, start=1):
        raw_headers.append(header_text)
        cell = ws_form.cell(row=5, column=col_idx, value=header_text)
        cell.font = HEADER_FONT
        cell.fill = SUBHEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = HEADER_BORDER
        col_letter = get_column_letter(col_idx)
        ws_form.column_dimensions[col_letter].width = col_width

    # Sample Data Rows (Rows 6 & 7) to guide the client
    sample_rows = [
        [
            1,
            "Taj Palace Hotel",
            "Hospitality",
            "Indian Hotels Company Ltd (IHCL)",
            "Lucknow, Uttar Pradesh",
            "Turnkey Interior Fit-out, Civil Refurbishment, MEP, Custom Joinery & Loose Furniture",
            "2,50,000 Sq. Ft.",
            "173 Executive Rooms, 19 Luxury Suites, 2 Grand Ballrooms, All-Day Dining Restaurant",
            "Completed (2024)",
            "A luxury heritage-inspired hotel interior seamlessly blending Awadhi craftsmanship with contemporary 5-star hospitality standards.",
            "The design brief focused on creating a palatial sanctuary in Lucknow. We integrated hand-carved jali panels, bespoke brass light fixtures, and rich marble flooring. High-performance acoustic paneling was installed in banquet halls.",
            "• Handcrafted Awadhi Jali Screens\n• Italian Botticino & Black Marquina Marble\n• Custom Acoustic Wall Paneling in Ballrooms\n• Automated Mood Lighting Integration",
            "Teak Wood Veneers, Italian Marble, Polished PVD Brass, Custom Silk Upholstery, Low-VOC Paints",
            "Lead Architect: BNP Design Studio | Lighting: Studio Glow",
            "\"BNP Interiors delivered exceptional craftsmanship and stuck to strict timelines for our flagship Lucknow hotel.\"",
            "https://drive.google.com/drive/folders/1example_taj_lucknow_photos"
        ],
        [
            2,
            "State Bank of India Corporate Tower",
            "Banking / Corporate",
            "State Bank of India",
            "Mumbai, Maharashtra",
            "Space Planning, Interior Architecture, Executive Suites & Auditorium Execution",
            "1,80,000 Sq. Ft.",
            "12 Floors Office Space, Boardrooms, 500-Seater Auditorium, Executive Lounge",
            "Completed (2023)",
            "Modern corporate headquarter design prioritizing employee wellness, ergonomic agility, and sustainable smart workplace automation.",
            "Designed for future-forward banking operations, featuring open-plan collaborative hubs, soundproof acoustic pods, and energy-efficient LED daylight harvesting systems.",
            "• Smart Glass Meeting Rooms\n• Biophilic Green Walls in Atrium\n• Ergonomic Workstations for 1200+ Staff\n• Centralized Building Automation",
            "Acoustic Fabric Panels, Engineered Oak Flooring, Recycled Aluminum Trim, Zero-VOC Carpets",
            "Lead Architect: BNP Commercial Team",
            "\"A seamless transformation of our corporate workspace. The team demonstrated high efficiency and professional rigor.\"",
            "https://drive.google.com/drive/folders/2example_sbi_tower_photos"
        ]
    ]

    for row_offset, row_data in enumerate(sample_rows, start=6):
        ws_form.row_dimensions[row_offset].height = 65
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws_form.cell(row=row_offset, column=col_idx, value=val)
            cell.font = SAMPLE_DATA_FONT
            cell.fill = ZEBRA_FILL
            cell.border = BORDER_ALL
            if col_idx in [1, 9]:
                cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
            elif col_idx == 16: # Drive link
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                cell.font = Font(name="Segoe UI", size=9, italic=True, color="2563EB", underline="single")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # Empty Data Entry Rows (Rows 8 to 25)
    for r in range(8, 26):
        ws_form.row_dimensions[r].height = 45
        for c in range(1, 17):
            cell = ws_form.cell(row=r, column=c)
            cell.font = DATA_FONT
            cell.border = BORDER_ALL
            cell.fill = WHITE_FILL
            if c == 1:
                cell.value = r - 7
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif c in [3, 9]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif c == 16:
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                cell.font = Font(name="Segoe UI", size=10, color="2563EB", underline="single")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # Data Validation (Dropdowns)
    cat_dv = DataValidation(type="list", formula1='"Hospitality, Commercial / Office, Banking, Residential, Retail, Healthcare, Institutional, Industrial, Other"', allow_blank=True)
    ws_form.add_data_validation(cat_dv)
    cat_dv.add("C6:C30")

    status_dv = DataValidation(type="list", formula1='"Completed, In Progress, Design Phase, Concept"', allow_blank=True)
    ws_form.add_data_validation(status_dv)
    status_dv.add("I6:I30")

    # ---------------------------------------------------------
    # SHEET 2: Instructions & Guidelines
    # ---------------------------------------------------------
    ws_guide = wb.create_sheet(title="Instructions_&_Guidelines")
    ws_guide.views.sheetView[0].showGridLines = True

    ws_guide.merge_cells("A1:E1")
    g_title = ws_guide["A1"]
    g_title.value = "GUIDELINES FOR SUBMITTING PROJECT CONTENT & IMAGES"
    g_title.font = TITLE_FONT
    g_title.fill = NAVY_HEADER_FILL
    g_title.alignment = Alignment(horizontal="center", vertical="center")
    ws_guide.row_dimensions[1].height = 35

    ws_guide.column_dimensions["A"].width = 6
    ws_guide.column_dimensions["B"].width = 28
    ws_guide.column_dimensions["C"].width = 50
    ws_guide.column_dimensions["D"].width = 50

    guide_headers = ["No.", "Topic / Section", "Instructions & Recommendations", "Example / Notes"]
    ws_guide.row_dimensions[3].height = 28
    for col_idx, text in enumerate(guide_headers, start=1):
        cell = ws_guide.cell(row=3, column=col_idx, value=text)
        cell.font = HEADER_FONT
        cell.fill = SUBHEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = HEADER_BORDER

    guidelines = [
        (
            1,
            "Google Drive Folder Link (Column P)",
            "1. Create a dedicated Google Drive / OneDrive folder for each project.\n2. Ensure folder sharing permission is set to 'Anyone with the link can view'.\n3. Paste the shareable URL in Column P (Drive Link of Image Folder).",
            "Example permissions: Restricted -> Change to 'Anyone with the link'"
        ),
        (
            2,
            "Photo Formatting & Resolution",
            "• Provide high-resolution JPEG / PNG photographs (min 1920x1080 px or higher).\n• Organize images inside the Drive folder into sub-folders if possible: '01_Hero_Featured', '02_Interiors', '03_Details', '04_Floor_Plans'.\n• Rename photos clearly (e.g., Lobby_01.jpg, Suite_Bed.jpg).",
            "Recommended photo count: 6 to 15 high quality shots per project."
        ),
        (
            3,
            "Project Description & Story",
            "• Short Summary: 2-3 sentences highlighting project significance.\n• Detailed Concept: Explain client brief, design inspiration, material selection, space distribution, and challenges overcome.",
            "Write in clear, professional tone suitable for prospective clients and award submissions."
        ),
        (
            4,
            "Key Features & Bullet Points",
            "List 4 to 8 key highlights that set this project apart (e.g. specialized ceiling acoustics, custom Italian marble inlay, bespoke lighting, LEED certification, etc.).",
            "Use bullet points or separate lines in the Excel cell."
        ),
        (
            5,
            "Confidentiality & Credits",
            "If any project details, client names, or financial figures are confidential, mark them as 'Confidential' or leave the optional fields blank. Always provide architectural and photography credits where applicable.",
            "e.g., Photography by XYZ Studios"
        )
    ]

    for idx, (num, topic, inst, ex) in enumerate(guidelines, start=4):
        ws_guide.row_dimensions[idx].height = 70
        
        c1 = ws_guide.cell(row=idx, column=1, value=num)
        c1.alignment = Alignment(horizontal="center", vertical="top")
        
        c2 = ws_guide.cell(row=idx, column=2, value=topic)
        c2.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        c2.font = Font(name="Segoe UI", size=10, bold=True, color="1F2937")
        
        c3 = ws_guide.cell(row=idx, column=3, value=inst)
        c3.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        
        c4 = ws_guide.cell(row=idx, column=4, value=ex)
        c4.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        c4.font = Font(name="Segoe UI", size=9, italic=True, color="4B5563")

        for c in range(1, 5):
            cell = ws_guide.cell(row=idx, column=c)
            cell.border = BORDER_ALL
            if idx % 2 == 0:
                cell.fill = ZEBRA_FILL
            else:
                cell.fill = WHITE_FILL

    # Output paths
    excel_path = r"c:\bnp resource\BNP_Interiors_Project_Content_Collection_Form.xlsx"
    csv_path = r"c:\bnp resource\BNP_Interiors_Project_Content_Collection_Form.csv"
    
    wb.save(excel_path)
    print(f"Excel created successfully at: {excel_path}")

    # Generate CSV as well
    with open(csv_path, mode="w", newline="", encoding="utf-8-sig") as csv_file:
        writer = csv.writer(csv_file)
        writer.writerow(raw_headers)
        for row in sample_rows:
            writer.writerow(row)
        for r in range(8, 26):
            empty_row = [r - 7] + [""] * 15
            writer.writerow(empty_row)
            
    print(f"CSV created successfully at: {csv_path}")

if __name__ == "__main__":
    create_project_data_sheet()
